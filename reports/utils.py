import datetime

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from api.models import (
    StudyGroup,
    Specialty
)
from api.serializers import (
    StudyGroupSerializer,
    SpecialtySerializer
)


def get_data() -> dict:
    specialties_data = []
    study_groups_data = []

    specialties: QuerySet = Specialty.objects. \
        select_related('curator').prefetch_related('disciplines')

    for specialty in specialties:
        specialties_data.append(SpecialtySerializer(instance=specialty).data)

    study_groups: QuerySet = StudyGroup.objects. \
        prefetch_related('students')

    for study_group in study_groups:

        study_groups_data.append(
            StudyGroupSerializer(instance=study_group).data
        )

    return {
        'specialties_data': specialties_data,
        'study_groups_data': study_groups_data
    }


def write_to_excel(specialties_data: list, study_groups_data: list) -> str:
    try:
        wb = Workbook()
        ws1 = wb.create_sheet('Specialties', 0)
        ws2 = wb.create_sheet('Study Groups', 1)

        headers1 = ['Name', 'Curator', 'Disciplines']
        headers2 = ['Group code', 'Students', 'Men', 'Women', 'Free slots']

        for col_num, header_title in enumerate(headers1, 1):
            col_letter = get_column_letter(col_num)
            ws1['{}1'.format(col_letter)] = header_title

        for col_num, header_title in enumerate(headers2, 1):
            col_letter = get_column_letter(col_num)
            ws2['{}1'.format(col_letter)] = header_title

        # Specialties page
        for row_num, obj in enumerate(specialties_data, 2):
            ws1.cell(row=row_num, column=1, value=obj['name'])

            name = 'name:' + obj['curator']['first_name']\
                   + obj['curator']['last_name']
            work_exp = 'work_exp: ' + str(obj['curator']['work_experience'])
            post = 'post: ' + obj['curator']['post']
            ws1.cell(
                row=row_num,
                column=2,
                value='\n'.join([name, work_exp, post])
            )

            disciplines = []
            for discipline in obj['disciplines']:
                disciplines.append(discipline['name'])

            ws1.cell(row=row_num, column=3, value='\n'.join(disciplines))

        # Study groups page
        for row_num, obj in enumerate(study_groups_data, 2):
            ws2.cell(row=row_num, column=1, value=obj['group_code'])

            students = []
            student_sex = {
                'male': 0,
                'female': 0,
            }
            student_cnt = 0
            for student in obj['students']:
                students.append(student['first_name'] + student['last_name'])
                student_sex[student['sex']] += 1
                student_cnt += 1

            ws2.cell(row=row_num, column=2, value='\n'.join(students))
            ws2.cell(row=row_num, column=3, value=student_sex['male'])
            ws2.cell(row=row_num, column=4, value=student_sex['female'])
            ws2.cell(
                row=row_num,
                column=5,
                value=obj['max_students'] - student_cnt
            )

        filename = 'media/' + \
                   str(datetime.datetime.now()) + '_report.xlsx'
        print()
        wb.save(filename)
        return filename
    except Exception:
        return ''
